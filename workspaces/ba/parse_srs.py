#!/usr/bin/env python3
"""Parse SRS xlsx — extract all 22 sheets into structured JSON for domain artifact generation."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SRS_PATH = Path("/home/hung/mvp-kho-bac/shared/specs/VDBAS_TT_SRS_III.1.1.2_TT.OUT.MANUAL_v7.xlsx")
OUTPUT_PATH = Path("/home/hung/mvp-kho-bac/workspaces/ba/srs_extract.json")

def sheet_to_rows(ws):
    """Convert worksheet to list of dicts keyed by header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = str(val).strip() if val is not None else ""
        data.append(d)
    return headers, data

def main():
    if not SRS_PATH.exists():
        print(f"ERROR: SRS not found at {SRS_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading: {SRS_PATH}")
    wb = openpyxl.load_workbook(SRS_PATH, read_only=True, data_only=True)

    result = {
        "source": SRS_PATH.name,
        "sheet_count": len(wb.sheetnames),
        "sheets": {}
    }

    for name in wb.sheetnames:
        ws = wb[name]
        headers, data = sheet_to_rows(ws)
        result["sheets"][name] = {
            "headers": headers,
            "row_count": len(data),
            "data": data
        }
        print(f"  Sheet '{name}': {len(headers)} cols, {len(data)} rows")

    wb.close()

    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nExtracted {result['sheet_count']} sheets → {OUTPUT_PATH}")
    print(f"Total rows: {sum(s['row_count'] for s in result['sheets'].values())}")

if __name__ == "__main__":
    main()
